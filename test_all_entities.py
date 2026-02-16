"""
Test document generation for all entity types using real trial balance data.
Generates documents for:
1. Foothills Conference Centre Pty Ltd (Company with COGS, depreciation, prior year)
2. Neri Family Trust (Trust with distributions, depreciation)
3. Mr Phillip North / Southy's Structures (Sole Trader with COGS, depreciation)
"""
import os
import sys
import django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
django.setup()

import openpyxl
from decimal import Decimal
from core.models import Client, Entity, FinancialYear, TrialBalanceLine, EntityOfficer, DepreciationAsset
from core.docgen import generate_financial_statements
from accounts.models import User
from datetime import date
import uuid


def parse_tb_excel(filepath):
    """Parse a trial balance Excel file and return list of account dicts."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]
    
    accounts = []
    current_section = None
    
    for row_idx in range(1, ws.max_row + 1):
        code = ws.cell(row=row_idx, column=1).value
        name = ws.cell(row=row_idx, column=2).value
        cy_dr = ws.cell(row=row_idx, column=3).value
        cy_cr = ws.cell(row=row_idx, column=4).value
        py_dr = ws.cell(row=row_idx, column=5).value
        py_cr = ws.cell(row=row_idx, column=6).value
        
        # Skip header rows
        if code and isinstance(code, str) and ("Trial Balance" in code or "ABN" in code 
            or code.startswith("$") or code.startswith("Trading") or "financial statements" in code.lower()):
            continue
        if name and isinstance(name, str) and name.startswith("$"):
            continue
        if code and isinstance(code, str) and code.strip() in ("2025", "2024"):
            continue
        
        # Section headers (no code, just text in col 1 or col 2)
        if code and isinstance(code, str) and not code.strip()[0:1].isdigit():
            section_name = code.strip()
            if section_name in ("Income", "Expenses", "Current Assets", "Non Current Assets", 
                              "Non-Current Assets", "Current Liabilities", "Non Current Liabilities",
                              "Non-Current Liabilities", "Equity", "Cost of Sales", "Net Profit"):
                current_section = section_name
            # Some section headers span into name column
            elif name is None and cy_dr is None:
                current_section = section_name
            continue
        
        # Skip Net Profit row
        if code and isinstance(code, str) and "Net Profit" in code:
            continue
        
        # Account line
        if code is not None:
            try:
                code_str = str(code).strip()
                # Handle codes like "0106" or "4005.01"
                if code_str and (code_str[0].isdigit()):
                    account = {
                        'code': code_str,
                        'name': str(name).strip() if name else '',
                        'cy_debit': Decimal(str(cy_dr)) if cy_dr else Decimal('0'),
                        'cy_credit': Decimal(str(cy_cr)) if cy_cr else Decimal('0'),
                        'py_debit': Decimal(str(py_dr)) if py_dr else Decimal('0'),
                        'py_credit': Decimal(str(py_cr)) if py_cr else Decimal('0'),
                        'section': current_section,
                    }
                    accounts.append(account)
            except (ValueError, IndexError):
                pass
        elif name is not None and cy_dr is None and cy_cr is None:
            # Might be a sub-section header with name in col 2
            pass
    
    wb.close()
    return accounts


def create_entity_and_import(entity_name, entity_type, abn, trading_as, filepath, 
                              officers=None, dep_assets=None, show_cents=False,
                              fy_end_date=date(2025, 6, 30)):
    """Create an entity, import trial balance, and generate document."""
    
    # Get or create user
    user = User.objects.first()
    
    # Create client
    client = Client.objects.create(
        name=entity_name.split(" Pty")[0].split(" Family")[0].split(" Mr ")[0] if " " in entity_name else entity_name,
        contact_email="test@example.com",
    )
    
    # Create entity
    entity = Entity.objects.create(
        client=client,
        entity_name=entity_name,
        entity_type=entity_type,
        abn=abn,
        trading_as=trading_as or '',
        show_cents=show_cents,
    )
    
    # Create officers
    if officers:
        for i, officer in enumerate(officers):
            EntityOfficer.objects.create(
                entity=entity,
                role=officer['role'],
                full_name=officer['name'],
                is_signatory=officer.get('signatory', True),
                display_order=i,
            )
    
    # Create financial year
    fy = FinancialYear.objects.create(
        entity=entity,
        year_label=f"FY{fy_end_date.year}",
        start_date=date(fy_end_date.year - 1, 7, 1),
        end_date=fy_end_date,
        status='draft',
    )
    
    # Parse and import trial balance
    accounts = parse_tb_excel(filepath)
    for acc in accounts:
        TrialBalanceLine.objects.create(
            financial_year=fy,
            account_code=acc['code'],
            account_name=acc['name'],
            debit=acc['cy_debit'],
            credit=acc['cy_credit'],
            prior_debit=acc['py_debit'],
            prior_credit=acc['py_credit'],
        )
    
    # Create depreciation assets if provided
    if dep_assets:
        for asset in dep_assets:
            DepreciationAsset.objects.create(
                financial_year=fy,
                **asset
            )
    
    print(f"  Created: {entity_name} ({entity_type})")
    print(f"  TB lines: {TrialBalanceLine.objects.filter(financial_year=fy).count()}")
    
    return entity, fy


def main():
    # Clean up previous test data
    Client.objects.all().delete()
    
    # =========================================================================
    # 1. FOOTHILLS CONFERENCE CENTRE PTY LTD (Company)
    # =========================================================================
    print("\n" + "="*60)
    print("  1. Foothills Conference Centre Pty Ltd (Company)")
    print("="*60)
    
    foothills_officers = [
        {'role': 'director', 'name': 'Anthony Caruso', 'signatory': True},
        {'role': 'director', 'name': 'Maria Caruso', 'signatory': True},
    ]
    
    # Sample depreciation assets for Foothills
    foothills_dep_assets = [
        # Buildings
        {'category': 'Buildings', 'asset_name': 'Conference Centre Building', 
         'method': 'prime_cost', 'rate': Decimal('2.5'),
         'opening_wdv': Decimal('450000'), 'depreciation_amount': Decimal('12500'),
         'closing_wdv': Decimal('437500'), 'total_cost': Decimal('500000'),
         'purchase_date': date(2015, 7, 1)},
        # Furniture & Fittings
        {'category': 'Furniture & Fittings', 'asset_name': 'Conference Room Furniture',
         'method': 'diminishing_value', 'rate': Decimal('20'),
         'opening_wdv': Decimal('15000'), 'depreciation_amount': Decimal('3000'),
         'closing_wdv': Decimal('12000'), 'total_cost': Decimal('25000'),
         'purchase_date': date(2020, 1, 15)},
        {'category': 'Furniture & Fittings', 'asset_name': 'Dining Room Tables & Chairs',
         'method': 'diminishing_value', 'rate': Decimal('20'),
         'opening_wdv': Decimal('8500'), 'depreciation_amount': Decimal('1700'),
         'closing_wdv': Decimal('6800'), 'total_cost': Decimal('18000'),
         'purchase_date': date(2019, 6, 1)},
        {'category': 'Furniture & Fittings', 'asset_name': 'Bar Equipment',
         'method': 'diminishing_value', 'rate': Decimal('20'),
         'opening_wdv': Decimal('3200'), 'depreciation_amount': Decimal('640'),
         'closing_wdv': Decimal('2560'), 'total_cost': Decimal('8000'),
         'purchase_date': date(2018, 3, 1)},
        # Motor Vehicles
        {'category': 'Motor Vehicles', 'asset_name': 'Toyota HiAce Van',
         'method': 'diminishing_value', 'rate': Decimal('25'),
         'opening_wdv': Decimal('22000'), 'depreciation_amount': Decimal('5500'),
         'closing_wdv': Decimal('16500'), 'total_cost': Decimal('45000'),
         'purchase_date': date(2021, 9, 1)},
        # Office Equipment
        {'category': 'Office Equipment', 'asset_name': 'Reception Computer System',
         'method': 'prime_cost', 'rate': Decimal('33.33'),
         'opening_wdv': Decimal('2000'), 'depreciation_amount': Decimal('1000'),
         'closing_wdv': Decimal('1000'), 'total_cost': Decimal('3000'),
         'purchase_date': date(2023, 1, 1)},
        {'category': 'Office Equipment', 'asset_name': 'Printer/Scanner',
         'method': 'prime_cost', 'rate': Decimal('33.33'),
         'opening_wdv': Decimal('800'), 'depreciation_amount': Decimal('400'),
         'closing_wdv': Decimal('400'), 'total_cost': Decimal('1200'),
         'purchase_date': date(2023, 7, 1)},
    ]
    
    entity1, fy1 = create_entity_and_import(
        entity_name="Foothills Conference Centre Pty Ltd",
        entity_type="company",
        abn="92 164 525 311",
        trading_as="",
        filepath="/home/ubuntu/upload/foothills.xlsx",
        officers=foothills_officers,
        dep_assets=foothills_dep_assets,
    )
    
    # Generate document
    output1 = "/home/ubuntu/mcs-platform/test_foothills_output.docx"
    buffer = generate_financial_statements(fy1.pk)
    with open(output1, 'wb') as f:
        f.write(buffer.read())
    print(f"  Document: {output1}")
    
    # =========================================================================
    # 2. NERI FAMILY TRUST (Trust)
    # =========================================================================
    print("\n" + "="*60)
    print("  2. Neri Family Trust (Trust)")
    print("="*60)
    
    neri_officers = [
        {'role': 'trustee', 'name': 'Claudio Neri', 'signatory': True},
        {'role': 'beneficiary', 'name': 'Claudio Neri', 'signatory': False},
        {'role': 'beneficiary', 'name': 'Sarah Neri', 'signatory': False},
    ]
    
    # Depreciation assets for Neri Trust
    neri_dep_assets = [
        {'category': 'Buildings', 'asset_name': 'Investment Property - Building Component',
         'method': 'prime_cost', 'rate': Decimal('2.5'),
         'opening_wdv': Decimal('296220'), 'depreciation_amount': Decimal('14811'),
         'closing_wdv': Decimal('281409'), 'total_cost': Decimal('370275'),
         'purchase_date': date(2015, 1, 1)},
        {'category': 'Office Equipment', 'asset_name': 'Office Equipment',
         'method': 'diminishing_value', 'rate': Decimal('40'),
         'opening_wdv': Decimal('2277'), 'depreciation_amount': Decimal('911'),
         'closing_wdv': Decimal('1366'), 'total_cost': Decimal('5000'),
         'purchase_date': date(2019, 7, 1)},
    ]
    
    entity2, fy2 = create_entity_and_import(
        entity_name="Neri Family Trust",
        entity_type="trust",
        abn="76 570 450 383",
        trading_as="",
        filepath="/home/ubuntu/upload/familytrust.xlsx",
        officers=neri_officers,
        dep_assets=neri_dep_assets,
        show_cents=True,
    )
    
    output2 = "/home/ubuntu/mcs-platform/test_neri_output.docx"
    buffer = generate_financial_statements(fy2.pk)
    with open(output2, 'wb') as f:
        f.write(buffer.read())
    print(f"  Document: {output2}")
    
    # =========================================================================
    # 3. MR PHILLIP NORTH / SOUTHY'S STRUCTURES (Sole Trader)
    # =========================================================================
    print("\n" + "="*60)
    print("  3. Mr Phillip North - Southy's Structures (Sole Trader)")
    print("="*60)
    
    north_officers = [
        {'role': 'proprietor', 'name': 'Phillip North', 'signatory': True},
    ]
    
    # Depreciation assets for sole trader
    north_dep_assets = [
        {'category': 'Office Equipment', 'asset_name': 'Office Equipment',
         'method': 'diminishing_value', 'rate': Decimal('40'),
         'opening_wdv': Decimal('0'), 'depreciation_amount': Decimal('0'),
         'closing_wdv': Decimal('0'), 'total_cost': Decimal('317.27'),
         'purchase_date': date(2018, 1, 1)},
        {'category': 'Motor Vehicles', 'asset_name': 'Work Vehicle',
         'method': 'diminishing_value', 'rate': Decimal('25'),
         'opening_wdv': Decimal('5846.45'), 'depreciation_amount': Decimal('1462'),
         'closing_wdv': Decimal('4384.45'), 'total_cost': Decimal('17074.45'),
         'purchase_date': date(2019, 7, 1)},
    ]
    
    entity3, fy3 = create_entity_and_import(
        entity_name="Mr Phillip North",
        entity_type="sole_trader",
        abn="77 157 285 382",
        trading_as="Southy's Structures",
        filepath="/home/ubuntu/upload/soletrader.xlsx",
        officers=north_officers,
        dep_assets=north_dep_assets,
        show_cents=True,
    )
    
    output3 = "/home/ubuntu/mcs-platform/test_soletrader_output.docx"
    buffer = generate_financial_statements(fy3.pk)
    with open(output3, 'wb') as f:
        f.write(buffer.read())
    print(f"  Document: {output3}")
    
    print("\n" + "="*60)
    print("  ALL DONE - 3 documents generated")
    print("="*60)


if __name__ == "__main__":
    main()
